"""
project_file.py — BioSafe Primer (.bsp) file handler.
Everything in memory. No disk writes ever.
Gel images stored as base64 inside project dict.
Excel/PDF generated into BytesIO buffers.
"""
import json, base64, copy, csv
from datetime import datetime
from io import BytesIO, StringIO
from collections import defaultdict

BSP_VERSION = "1.0"


def new_project(name, vector_name, vector_length, vector_sequence, vector_features):
    return {
        "bsp_version":      BSP_VERSION,
        "project_name":     name,
        "vector_name":      vector_name,
        "vector_length":    vector_length,
        "vector_sequence":  vector_sequence,
        "vector_features":  vector_features,
        "created_at":       datetime.now().strftime("%Y-%m-%d %H:%M"),
        "last_saved":       None,
        "primers":          [],
        "pcr_runs":         [],
        "redesign_history": [],
    }


def save_project_bytes(state):
    s = copy.deepcopy(state)
    s["last_saved"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    return json.dumps(s, indent=2, ensure_ascii=False).encode("utf-8")


def load_project_bytes(file_bytes):
    state = json.loads(file_bytes.decode("utf-8"))
    if "project_name" not in state:
        raise ValueError("Not a valid BioSafe Primer (.bsp) file.")
    return state


def encode_gel_image(file_bytes):
    return base64.b64encode(file_bytes).decode("utf-8")


def decode_gel_image(b64_string):
    return base64.b64decode(b64_string) if b64_string else None


def get_best_primers(state):
    seen = {}
    for p in state.get("primers", []):
        an = p["amplicon_num"]
        if an not in seen or p.get("version",1) > seen[an].get("version",1):
            seen[an] = p
    return sorted(seen.values(), key=lambda x: x["amplicon_num"])


def add_primers(state, primers):
    state["primers"].extend(primers)


def update_primer_status(state, primer_id, new_status):
    for p in state["primers"]:
        if p.get("_id") == primer_id:
            p["status"] = new_status
            return


def update_amplicon_name(state, primer_id, new_name):
    amp_num = None
    for p in state["primers"]:
        if p.get("_id") == primer_id:
            amp_num = p["amplicon_num"]
            break
    if amp_num is not None:
        for p in state["primers"]:
            if p["amplicon_num"] == amp_num:
                p["amplicon_name"] = new_name


def add_pcr_run(state, primer_id, result, gel_b64, lane_number,
                notes, amplicon_num, fp_sequence, rp_sequence):
    state["pcr_runs"].append({
        "id":            len(state["pcr_runs"]) + 1,
        "primer_id":     primer_id,
        "amplicon_num":  amplicon_num,
        "fp_sequence":   fp_sequence,
        "rp_sequence":   rp_sequence,
        "result":        result,
        "gel_image_b64": gel_b64,
        "lane_number":   lane_number,
        "notes":         notes,
        "run_date":      datetime.now().strftime("%Y-%m-%d %H:%M"),
    })


def add_redesign_history(state, amplicon_num, old_primer_id,
                          ext_left, ext_right, reason, failure_type,
                          attempt_num, upstream_overlap, downstream_overlap):
    state["redesign_history"].append({
        "amplicon_num":               amplicon_num,
        "old_primer_id":              old_primer_id,
        "extension_left":             ext_left,
        "extension_right":            ext_right,
        "failure_type":               failure_type,
        "attempt_num":                attempt_num,
        "upstream_overlap_result":    upstream_overlap,
        "downstream_overlap_result":  downstream_overlap,
        "reason":                     reason,
        "redesign_date":              datetime.now().strftime("%Y-%m-%d %H:%M"),
    })


def assign_ids(primers):
    for i, p in enumerate(primers):
        p["_id"] = i
    return primers


def get_project_stats(state):
    best  = get_best_primers(state)
    total = len(best)
    done  = sum(1 for p in best if p.get("status") == "Done")
    return {"total": total, "done": done}


def primers_to_excel_bytes(primers, project_name):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    STATUS_COLORS = {
        'Pending':'FFF9C4','Done':'C8E6C9','Failed':'FFCDD2',
        'Overlap Violation':'E1BEE7','Design Failed':'F8BBD9','Redesigned':'B3E5FC',
    }
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Primers"

    total_cols = 22
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    tc = ws['A1']
    tc.value = f"Overlapping PCR Primer Design — {project_name}"
    tc.font  = Font(bold=True, size=13, color='FFFFFF')
    tc.fill  = PatternFill('solid', fgColor='1A237E')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = [
        'Amp #','Amplicon Name','Ver','Status',
        "FP Sequence (5'→3')",'FP\nLen','FP\nTm(°C)','FP\nGC%',
        'FP Hairpin\nTm(°C)',"FP 3'Stab\n(ΔG)",'FP\nPenalty',
        "RP Sequence (5'→3')",'RP\nLen','RP\nTm(°C)','RP\nGC%',
        'RP Hairpin\nTm(°C)',"RP 3'Stab\n(ΔG)",'RP\nPenalty',
        'Pair\nPenalty','Amp\nLen(bp)','Overlap\nUpstream(bp)','Overlap\nDownstream(bp)'
    ]
    hfill  = PatternFill('solid', fgColor='283593')
    hfont  = Font(bold=True, color='FFFFFF', size=9)
    border = Border(
        left=Side(style='thin',color='9FA8DA'), right=Side(style='thin',color='9FA8DA'),
        top=Side(style='thin',color='9FA8DA'), bottom=Side(style='thin',color='9FA8DA')
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill=hfill; cell.font=hfont; cell.border=border
        cell.alignment=Alignment(horizontal='center',wrap_text=True)
    ws.row_dimensions[2].height = 38

    for ri, p in enumerate(primers, 3):
        status = p.get('status','Pending')
        rfill  = PatternFill('solid',fgColor=STATUS_COLORS.get(status,'FFFFFF'))
        row = [
            p['amplicon_num'],
            p.get('amplicon_name', f"Amplicon_{p['amplicon_num']}"),
            p.get('version',1), status,
            p['fp_sequence'],p['fp_length'],p['fp_tm'],p['fp_gc'],
            p.get('fp_hairpin_tm',0),p.get('fp_end_stability',0),p.get('fp_penalty',0),
            p['rp_sequence'],p['rp_length'],p['rp_tm'],p['rp_gc'],
            p.get('rp_hairpin_tm',0),p.get('rp_end_stability',0),p.get('rp_penalty',0),
            p.get('pair_penalty',0),p['amplicon_length'],
            p.get('overlap_prev') if p.get('overlap_prev') is not None else 'N/A',
            p.get('overlap_next') if p.get('overlap_next') is not None else 'N/A',
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill=rfill; cell.border=border
            cell.alignment=Alignment(horizontal='center',wrap_text=True)
            if ci in [5,12]:
                cell.font=Font(name='Courier New',size=8)
                cell.alignment=Alignment(horizontal='left')

    col_widths=[7,18,5,12,36,7,8,7,10,10,9,36,7,8,7,10,10,9,9,9,12,13]
    for i,w in enumerate(col_widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes='A3'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def primers_to_summary_excel_bytes(primers, project_name):
    """
    SHORT-FORMAT result export (.xlsx).
    Columns: Amplicon No, Amplicon Name, FP_SEQUENCE, fp_gc, fp_tm,
             RP_SEQUENCE, rp_gc, rp_tm, amplicon_length,
             overlap_prev, overlap_next
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    headers = ['Amplicon No', 'Amplicon Name', 'FP_SEQUENCE', 'fp_gc', 'fp_tm',
               'RP_SEQUENCE', 'rp_gc', 'rp_tm', 'amplicon_length',
               'overlap_prev', 'overlap_next']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Primer Summary"

    total_cols = len(headers)
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    tc = ws['A1']
    tc.value = f"Primer Summary (Short Format) — {project_name}"
    tc.font  = Font(bold=True, size=13, color='FFFFFF')
    tc.fill  = PatternFill('solid', fgColor='1A237E')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    hfill  = PatternFill('solid', fgColor='283593')
    hfont  = Font(bold=True, color='FFFFFF', size=10)
    border = Border(
        left=Side(style='thin', color='9FA8DA'), right=Side(style='thin', color='9FA8DA'),
        top=Side(style='thin', color='9FA8DA'), bottom=Side(style='thin', color='9FA8DA')
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = hfill; cell.font = hfont; cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[2].height = 22

    for ri, p in enumerate(primers, 3):
        row = [
            p['amplicon_num'],
            p.get('amplicon_name', f"Amplicon_{p['amplicon_num']}"),
            p['fp_sequence'], p['fp_gc'], p['fp_tm'],
            p['rp_sequence'], p['rp_gc'], p['rp_tm'],
            p['amplicon_length'],
            p.get('overlap_prev') if p.get('overlap_prev') is not None else 'N/A',
            p.get('overlap_next') if p.get('overlap_next') is not None else 'N/A',
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            if ci in [3, 6]:
                cell.font = Font(name='Courier New', size=9)
                cell.alignment = Alignment(horizontal='left')

    col_widths = [10, 20, 30, 8, 8, 30, 8, 8, 13, 12, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def primers_to_order_csv_bytes(primers):
    """
    "Order Primers" export (.csv) — one row per FP and per RP, ready to
    paste into an oligo-order sheet.
    Columns: Primer Name, Sequence (5'->3'), Number of Bases
    """
    sio = StringIO()
    writer = csv.writer(sio)
    writer.writerow(["Primer Name", "Sequence (5'->3')", "Number of Bases"])
    for p in primers:
        if p.get('fp_sequence') == 'DESIGN_FAILED':
            continue
        label = p.get('amplicon_name') or f"Amplicon_{p['amplicon_num']}"
        writer.writerow([f"{label}_FP", p['fp_sequence'], p['fp_length']])
        writer.writerow([f"{label}_RP", p['rp_sequence'], p['rp_length']])
    buf = BytesIO(sio.getvalue().encode('utf-8'))
    buf.seek(0)
    return buf


def primers_to_pdf_bytes(project_name, primers, pcr_runs):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, Image as RLImage, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf  = BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=landscape(A4),
                              leftMargin=1.5*cm, rightMargin=1.5*cm,
                              topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('t',parent=styles['Title'],fontSize=15,
                              textColor=colors.HexColor('#1A237E'),spaceAfter=6)
    sec_s   = ParagraphStyle('s',parent=styles['Heading2'],fontSize=11,
                              textColor=colors.HexColor('#283593'),spaceAfter=5)
    body_s  = ParagraphStyle('b',parent=styles['Normal'],fontSize=8,spaceAfter=3)
    story   = []

    story.append(Paragraph(f"Overlapping PCR Primer Report: {project_name}", title_s))
    story.append(Paragraph(
        "BioSafe Primer | Division of Plant Physiology, ICAR-IARI", body_s))
    story.append(Spacer(1,0.3*cm))
    story.append(PageBreak())

    story.append(Paragraph("Primer Design Summary", sec_s))
    thead = [['Amp#','Name','Ver','Status',
              'Forward Primer','FP\nLen','FP\nTm','FP\nGC%',
              'FP\nHairpin',"FP\n3'Stab",'FP\nPenalty',
              'Reverse Primer','RP\nLen','RP\nTm','RP\nGC%',
              'RP\nHairpin',"RP\n3'Stab",'RP\nPenalty',
              'Pair\nPenalty','Amp\nLen','Overlap\nUp','Overlap\nDown']]
    for p in primers:
        prev_ov = p.get('overlap_prev')
        next_ov = p.get('overlap_next')
        thead.append([
            str(p['amplicon_num']),
            p.get('amplicon_name',f"Amplicon_{p['amplicon_num']}"),
            str(p.get('version',1)), p.get('status','Pending'),
            p['fp_sequence'],str(p['fp_length']),f"{p['fp_tm']}°C",f"{p['fp_gc']}%",
            f"{p.get('fp_hairpin_tm',0)}°C",f"{p.get('fp_end_stability',0)}",
            str(p.get('fp_penalty',0)),
            p['rp_sequence'],str(p['rp_length']),f"{p['rp_tm']}°C",f"{p['rp_gc']}%",
            f"{p.get('rp_hairpin_tm',0)}°C",f"{p.get('rp_end_stability',0)}",
            str(p.get('rp_penalty',0)),str(p.get('pair_penalty',0)),
            str(p['amplicon_length']),
            str(prev_ov) if prev_ov is not None else 'N/A',
            str(next_ov) if next_ov is not None else 'N/A',
        ])
    t = Table(thead, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1A237E')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),5.5),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),
         [colors.HexColor('#F5F5FF'),colors.HexColor('#E8EAF6')]),
        ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#9FA8DA')),
        ('FONTNAME',(4,1),(4,-1),'Courier'),
        ('FONTNAME',(11,1),(11,-1),'Courier'),
    ]))
    story.append(t)

    if pcr_runs:
        story.append(PageBreak())
        story.append(Paragraph("PCR Run Log", sec_s))
        rhead = [['Run Date','Amplicon #','Result','Lane','Notes']]
        for r in pcr_runs:
            rhead.append([r.get('run_date',''),str(r.get('amplicon_num','')),
                          r.get('result',''),str(r.get('lane_number','')),r.get('notes','')])
        rt = Table(rhead, repeatRows=1)
        rt.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#283593')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTSIZE',(0,0),(-1,-1),7),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#9FA8DA')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),
             [colors.HexColor('#FFF9C4'),colors.white]),
        ]))
        story.append(rt)

        runs_with_gel = [r for r in pcr_runs if r.get('gel_image_b64')]
        if runs_with_gel:
            story.append(PageBreak())
            story.append(Paragraph("Gel Images", sec_s))
            amp_gels = defaultdict(list)
            for r in runs_with_gel:
                amp_gels[r.get('amplicon_num','?')].append(r)
            for amp_num in sorted(amp_gels.keys()):
                story.append(Paragraph(f"Amplicon {amp_num}", sec_s))
                gel_row = []
                for r in amp_gels[amp_num]:
                    img_bytes = decode_gel_image(r['gel_image_b64'])
                    img_buf   = BytesIO(img_bytes)
                    rl        = '✅ Pass' if r.get('result')=='Pass' else '❌ Fail'
                    rc        = (colors.HexColor('#1b5e20')
                                 if r.get('result')=='Pass'
                                 else colors.HexColor('#b71c1c'))
                    try:
                        img_el = RLImage(img_buf, width=7*cm, height=5*cm)
                    except Exception:
                        continue  # skip invalid/corrupt image
                    cap       = Paragraph(
                        f"<b>{rl}</b> Lane {r.get('lane_number','')} {r.get('run_date','')}",
                        ParagraphStyle('gc',parent=body_s,fontSize=7,
                                       textColor=rc,alignment=TA_CENTER))
                    gel_row.append([img_el, cap])
                    if len(gel_row) == 3:
                        tbl = Table([[item for pair in gel_row for item in pair]],
                                    colWidths=[7*cm,3.5*cm]*3)
                        tbl.setStyle(TableStyle([
                            ('VALIGN',(0,0),(-1,-1),'TOP'),
                            ('ALIGN',(0,0),(-1,-1),'CENTER'),
                            ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#9FA8DA')),
                        ]))
                        story.append(tbl)
                        story.append(Spacer(1,0.2*cm))
                        gel_row = []
                if gel_row:
                    tbl = Table([[item for pair in gel_row for item in pair]],
                                colWidths=[7*cm,3.5*cm]*len(gel_row))
                    tbl.setStyle(TableStyle([
                        ('VALIGN',(0,0),(-1,-1),'TOP'),
                        ('ALIGN',(0,0),(-1,-1),'CENTER'),
                        ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#9FA8DA')),
                    ]))
                    story.append(tbl)
                    story.append(Spacer(1,0.3*cm))

    doc.build(story)
    buf.seek(0)
    return buf
