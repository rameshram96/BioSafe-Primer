import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                  Paragraph, Spacer, Image as RLImage, PageBreak)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

STATUS_COLORS = {
    'Pending':           'FFF9C4',
    'Done':              'C8E6C9',
    'Failed':            'FFCDD2',
    'Overlap Violation': 'E1BEE7',
    'Design Failed':     'F8BBD9',
    'Redesigned':        'B3E5FC',
}


def export_primers_excel(primers, project_name, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Primers"

    # Title row
    total_cols = 22
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    tc = ws['A1']
    tc.value = f"Overlapping PCR Primer Design — {project_name}"
    tc.font  = Font(bold=True, size=13, color='FFFFFF')
    tc.fill  = PatternFill('solid', fgColor='1A237E')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = [
        'Amp #', 'Amplicon Name', 'Ver', 'Status',
        "FP Sequence (5'→3')", 'FP\nLen', 'FP\nTm(°C)', 'FP\nGC%',
        'FP Hairpin\nTm(°C)', 'FP 3\' Stab\n(ΔG)', 'FP\nPenalty',
        "RP Sequence (5'→3')", 'RP\nLen', 'RP\nTm(°C)', 'RP\nGC%',
        'RP Hairpin\nTm(°C)', 'RP 3\' Stab\n(ΔG)', 'RP\nPenalty',
        'Pair\nPenalty',
        'Amp\nLen(bp)', 'Overlap\nUpstream(bp)', 'Overlap\nDownstream(bp)'
    ]

    hfill = PatternFill('solid', fgColor='283593')
    hfont = Font(bold=True, color='FFFFFF', size=9)
    border = Border(
        left=Side(style='thin', color='9FA8DA'),
        right=Side(style='thin', color='9FA8DA'),
        top=Side(style='thin', color='9FA8DA'),
        bottom=Side(style='thin', color='9FA8DA')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = hfill; cell.font = hfont; cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[2].height = 38

    for ri, p in enumerate(primers, 3):
        status  = p.get('status', 'Pending')
        rfill   = PatternFill('solid', fgColor=STATUS_COLORS.get(status, 'FFFFFF'))
        prev_ov = p.get('overlap_prev')
        next_ov = p.get('overlap_next')
        row = [
            p['amplicon_num'],
            p.get('amplicon_name', f"Amplicon_{p['amplicon_num']}"),
            p.get('version', 1), status,
            p['fp_sequence'], p['fp_length'], p['fp_tm'], p['fp_gc'],
            p.get('fp_hairpin_tm', 0), p.get('fp_end_stability', 0),
            p.get('fp_penalty', 0),
            p['rp_sequence'], p['rp_length'], p['rp_tm'], p['rp_gc'],
            p.get('rp_hairpin_tm', 0), p.get('rp_end_stability', 0),
            p.get('rp_penalty', 0), p.get('pair_penalty', 0),
            p['amplicon_length'],
            prev_ov if prev_ov is not None else 'N/A',
            next_ov if next_ov is not None else 'N/A',
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = rfill; cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            if ci in [4, 11]:  # sequence columns
                cell.font = Font(name='Courier New', size=8)
                cell.alignment = Alignment(horizontal='left')

    col_widths = [7,18,5,12, 36,7,8,7, 10,10,9,
                  36,7,8,7, 10,10,9,9,
                  9,12,13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A3'
    wb.save(output_path)
    return output_path


def export_full_report_pdf(project_name, primers, pcr_runs,
                            linear_map_path, output_path):
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('t', parent=styles['Title'], fontSize=15,
                              textColor=colors.HexColor('#1A237E'), spaceAfter=6)
    sec_s   = ParagraphStyle('s', parent=styles['Heading2'], fontSize=11,
                              textColor=colors.HexColor('#283593'), spaceAfter=5)
    body_s  = ParagraphStyle('b', parent=styles['Normal'], fontSize=8, spaceAfter=3)
    story   = []

    story.append(Paragraph(f"Overlapping PCR Primer Report: {project_name}", title_s))
    story.append(Paragraph(
        "Generated for GMO regulatory exemption — full vector coverage verification.",
        body_s))
    story.append(Spacer(1, 0.3*cm))

    # Linear map
    if linear_map_path and os.path.exists(linear_map_path):
        story.append(Paragraph("Linear Vector Map with Primer Positions", sec_s))
        story.append(RLImage(linear_map_path, width=24*cm, height=7*cm))
        story.append(Spacer(1, 0.3*cm))

    story.append(PageBreak())

    # Primer table
    story.append(Paragraph("Primer Design Summary", sec_s))
    thead = [['Amp#','Amplicon Name','Ver','Status',
              'Forward Primer','FP\nLen','FP\nTm','FP\nGC%',
              'FP\nHairpin','FP\n3\'Stab','FP\nPenalty',
              'Reverse Primer','RP\nLen','RP\nTm','RP\nGC%',
              'RP\nHairpin','RP\n3\'Stab','RP\nPenalty',
              'Pair\nPenalty','Amp\nLen',
              'Overlap\nUp','Overlap\nDown']]

    for p in primers:
        prev_ov = p.get('overlap_prev')
        next_ov = p.get('overlap_next')
        thead.append([
            str(p['amplicon_num']),
            p.get('amplicon_name', f"Amplicon_{p['amplicon_num']}"),
            str(p.get('version',1)),
            p.get('status','Pending'),
            p['fp_sequence'], str(p['fp_length']),
            f"{p['fp_tm']}°C", f"{p['fp_gc']}%",
            f"{p.get('fp_hairpin_tm',0)}°C",
            f"{p.get('fp_end_stability',0)}",
            str(p.get('fp_penalty',0)),
            p['rp_sequence'], str(p['rp_length']),
            f"{p['rp_tm']}°C", f"{p['rp_gc']}%",
            f"{p.get('rp_hairpin_tm',0)}°C",
            f"{p.get('rp_end_stability',0)}",
            str(p.get('rp_penalty',0)),
            str(p.get('pair_penalty',0)),
            str(p['amplicon_length']),
            str(prev_ov) if prev_ov is not None else 'N/A',
            str(next_ov) if next_ov is not None else 'N/A',
        ])

    t = Table(thead, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A237E')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 6),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1),
         [colors.HexColor('#F5F5FF'), colors.HexColor('#E8EAF6')]),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#9FA8DA')),
        ('FONTNAME',   (3,1), (3,-1), 'Courier'),
        ('FONTNAME',   (10,1),(10,-1),'Courier'),
    ]))
    story.append(t)

    # PCR run log + gel images
    if pcr_runs:
        story.append(PageBreak())
        story.append(Paragraph("PCR Run Log", sec_s))

        # Summary table first
        rhead = [['Run Date','Amplicon #','Result','Lane','Primers Used','Notes']]
        for r in pcr_runs:
            rhead.append([
                r.get('run_date',''),
                str(r.get('amplicon_num','')),
                r.get('result',''),
                str(r.get('lane_number','')),
                f"FP: {str(r.get('fp_sequence','—'))[:20]}…\nRP: {str(r.get('rp_sequence','—'))[:20]}…",
                r.get('notes','')
            ])
        rt = Table(rhead, repeatRows=1,
                   colWidths=[2.5*cm, 2*cm, 1.8*cm, 1.5*cm, 7*cm, 5.5*cm])
        rt.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#283593')),
            ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 7),
            ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME',      (4,1), (4,-1), 'Courier'),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#9FA8DA')),
            ('ROWBACKGROUNDS',(0,1), (-1,-1),
             [colors.HexColor('#FFF9C4'), colors.white]),
        ]))
        story.append(rt)

        # Gel images — grouped per amplicon
        runs_with_gel = [r for r in pcr_runs
                         if r.get('gel_image_path') and
                            os.path.exists(r.get('gel_image_path',''))]

        if runs_with_gel:
            story.append(PageBreak())
            story.append(Paragraph("Gel Images", sec_s))
            story.append(Spacer(1, 0.2*cm))

            # Group by amplicon number
            from collections import defaultdict
            amp_gels = defaultdict(list)
            for r in runs_with_gel:
                amp_gels[r.get('amplicon_num', '?')].append(r)

            for amp_num in sorted(amp_gels.keys()):
                runs_for_amp = amp_gels[amp_num]
                story.append(Paragraph(
                    f"Amplicon {amp_num}", sec_s
                ))

                # Lay out up to 3 gel images per row
                gel_row = []
                for r in runs_for_amp:
                    gp     = r.get('gel_image_path')
                    result = r.get('result', '')
                    date   = r.get('run_date', '')
                    lane   = r.get('lane_number', '')
                    notes  = r.get('notes', '')

                    result_color = (colors.HexColor('#1b5e20')
                                    if result == 'Pass'
                                    else colors.HexColor('#b71c1c'))
                    result_label = '✅ Pass' if result == 'Pass' else '❌ Fail'

                    img_cell = RLImage(gp, width=7*cm, height=5*cm)
                    caption  = Paragraph(
                        f"<b>{result_label}</b>  Lane {lane}  {date}<br/>"
                        f"<font size=6>{notes[:60] if notes else ''}</font>",
                        ParagraphStyle('gc', parent=body_s, fontSize=7,
                                       textColor=result_color,
                                       alignment=TA_CENTER)
                    )
                    gel_row.append([img_cell, caption])

                    if len(gel_row) == 3:
                        # Render row of 3
                        flat = []
                        for img_c, cap_c in gel_row:
                            flat.append([img_c, cap_c])
                        tbl = Table([[item for pair in gel_row for item in pair]],
                                    colWidths=[7*cm, 3.5*cm]*len(gel_row))
                        tbl.setStyle(TableStyle([
                            ('VALIGN',  (0,0),(-1,-1),'TOP'),
                            ('ALIGN',   (0,0),(-1,-1),'CENTER'),
                            ('GRID',    (0,0),(-1,-1),0.3,
                             colors.HexColor('#9FA8DA')),
                            ('PADDING', (0,0),(-1,-1),4),
                        ]))
                        story.append(tbl)
                        story.append(Spacer(1, 0.2*cm))
                        gel_row = []

                # Remaining images (< 3 in last row)
                if gel_row:
                    tbl = Table([[item for pair in gel_row for item in pair]],
                                colWidths=[7*cm, 3.5*cm]*len(gel_row))
                    tbl.setStyle(TableStyle([
                        ('VALIGN',  (0,0),(-1,-1),'TOP'),
                        ('ALIGN',   (0,0),(-1,-1),'CENTER'),
                        ('GRID',    (0,0),(-1,-1),0.3,
                         colors.HexColor('#9FA8DA')),
                        ('PADDING', (0,0),(-1,-1),4),
                    ]))
                    story.append(tbl)
                    story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    return output_path
